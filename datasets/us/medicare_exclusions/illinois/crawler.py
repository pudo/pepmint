from base64 import b64decode
import json
from typing import Dict, Generator
from rigour.mime.types import XLSX
from openpyxl import load_workbook
from normality import slugify, stringify
from datetime import datetime
import openpyxl

from zavod import Context, helpers as h
from zavod.shed.zyte_api import fetch_raw


def parse_sheet(
    context: Context,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    skiprows: int = 0,
) -> Generator[dict, None, None]:
    headers = None

    row_counter = 0

    for row in sheet.iter_rows():
        # Increment row counter
        row_counter += 1

        # Skip the desired number of rows
        if row_counter <= skiprows:
            continue
        cells = [c.value for c in row]
        if headers is None:
            headers = []
            for idx, cell in enumerate(cells):
                if cell is None:
                    cell = f"column_{idx}"
                headers.append(slugify(cell))
            continue

        record = {}
        for header, value in zip(headers, cells):
            if isinstance(value, datetime):
                value = value.date()
            record[header] = stringify(value)
        if len(record) == 0:
            continue
        yield record


def crawl_item(row: Dict[str, str], context: Context, is_excluded: bool):

    name = row.pop("provider-name")

    if not name:
        return

    entity = context.make("LegalEntity")
    entity.id = context.make_id(name)
    entity.add("name", name)
    entity.add("alias", row.pop("doing-business-as-name"))
    if row.get("npi") != "N/A":
        entity.add("notes", "NPI: " + row.pop("npi"))
    else:
        row.pop("npi")
    if is_excluded:
        entity.add("topics", "debarment")

    sanction = h.make_sanction(context, entity)
    sanction.add("startDate", row.pop("termination-effective-date"))
    sanction.add("reason", row.pop("termination-authority"))

    if row.get("reinstatement-effective-date"):
        sanction.add("endDate", row.pop("reinstatement-effective-date"))

    context.emit(entity, target=is_excluded)
    context.emit(sanction)

    context.audit_data(row)


def crawl_excel_url(context: Context):
    doc = context.fetch_html(context.data_url)
    return doc.find(".//*[@about='/provider-termination']").find(".//a").get("href")


def crawl(context: Context) -> None:
    zyte_data = {
        "url": context.data_url,
        "browserHtml": True,
        "networkCapture": [
            {
                "filterType": "url",
                "httpResponseBody": True,
                "value": "download-as-excel",
                "matchType": "contains",
            },
        ],
        "actions": [
            {
                "action": "click",
                "selector": {
                    "type": "css",
                    "value": ".entitylist-download",
                },
            },
            {
                "action": "waitForResponse",
                "urlPattern": "download-as-excel",
                "urlMatchingOptions": "contains", 
                "timeout": 15,
            }
        ],
    }

    response = fetch_raw(context, zyte_data)
    for action in response["actions"]:
        context.log.info("action", action)
    assert len(response["networkCapture"]) == 1, response["networkCapture"]

    key_response = b64decode(response["networkCapture"][0]["httpResponseBody"]).decode("utf-8")
    key = json.loads(key_response)["sessionKey"]

    url = response["networkCapture"][0]["url"] + "?key=" + key
    context.fetch_resource("source.xlsx", url)
    #context.export_resource(path, XLSX, title=context.SOURCE_TITLE)


#
# wb = load_workbook(path, read_only=True)
#
## Currently terminated providers
# for item in parse_sheet(context, wb["Termination List"]):
#    crawl_item(item, context, True)
#
## Providers that where terminated but are now reinstated
# for item in parse_sheet(context, wb["Reinstated Providers"]):
#    crawl_item(item, context, False)
